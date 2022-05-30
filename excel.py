from calendar import c
from ctypes import alignment
from curses import A_ALTCHARSET
from re import A
import openpyxl
from dados_acao import acao
from dados_moeda import moeda
from random import randint
from openpyxl.drawing.image import Image
from openpyxl.styles.alignment import Alignment

""" url = 'https://gustavocampanha.github.io/carteira_investimento/index.html' """

""" ls_acao = acao(url)"""
""" ls_moeda = moeda(url) """


ls_acao = [{'Nome da Ação': 'EMBR3.SA', 'Quantidade de Ações': 100, 'Valor da Ação': 12.5, 'Valor Investido': 1250.0}, {'Nome da Ação': 'PRIO3.SA', 'Quantidade de Ações': 325, 'Valor da Ação': 27.98, 'Valor Investido': 9093.5}]

ls_moeda = [{'Nome da Moeda': 'BRL=X', 'Quantidade de Moeda': 3000, 'Valor da Moeda': 4.7495, 'Valor Investido': 14248.5}, {'Nome da Moeda': 'CNYBRL=X', 'Quantidade de Moeda': 20000, 'Valor da Moeda': 0.70916, 'Valor Investido': 14183.2}]

#Adicionando o Tipo de Investimento no Dicionário de cada Ativo
for a in range(0, len(ls_acao)):
    ls_acao[a]['Tipo'] = 'Ação'

for a in range(0, len(ls_moeda)):
    ls_moeda[a]['Tipo'] = 'Moeda'


#Criar uma planilha chamada book
arquivo = openpyxl.Workbook()
ws = arquivo.active
ws.title = "Tabela"

#Descobrindo o patrimonio total da pessoa
valor_da_carteira = 0
for v_acao in range(0, len(ls_acao)):
    valor_da_carteira += ls_acao[v_acao]['Valor Investido']
for v_moeda in range(0, len(ls_moeda)):
    valor_da_carteira += ls_moeda[v_moeda]['Valor Investido']


#Criando outras páginas
arquivo.create_sheet('Grafico 1')
arquivo.create_sheet('Grafico 2')
arquivo.create_sheet('Grafico 3')
arquivo.create_sheet('Valor da Carteira')

#Selecionar uma página
pag_inicial = arquivo['Tabela']

#Criando a lista do header
header = ["Nome", "Quantidade", "Valor Unitário", "Valor Investido", "Tipo"]

#Adicionando o header na tabela
for col in range(0, 5):
    pag_inicial.cell(row=2,column=col+2).value = header[col]

tam_acao = len(ls_acao)
tam_moeda = len(ls_moeda)
tam_total = tam_acao + tam_moeda

#Adicionando os nomes das ações na coluna 2
for lin in range(0, tam_acao):
    pag_inicial.cell(row=lin+3, column=2).value = ls_acao[lin]['Nome da Ação']

#Adicionando as quantidades das ações na coluna 3
for lin in range(0, tam_acao):
    pag_inicial.cell(row=lin+3, column=3).value = ls_acao[lin]['Quantidade de Ações']

#Adicionando o valor da ação na coluna 4
for lin in range(0, tam_acao):
    pag_inicial.cell(row=lin+3, column=4).value = ls_acao[lin]['Valor da Ação']

#Adicionando o valor investido no ativo na coluna 5
for lin in range(0, tam_acao):
    pag_inicial.cell(row=lin+3, column=5).value = ls_acao[lin]['Valor Investido']

#Adicionando o tipo de ativo ação na coluna 6
for lin in range(0, tam_acao):
    pag_inicial.cell(row=lin+3, column=6).value = ls_acao[lin]['Tipo']

#Adicionando os nomes das moedas na coluna 2
for lin in range(0, tam_moeda):
    pag_inicial.cell(row=lin+tam_acao+3, column=2).value = ls_moeda[lin]['Nome da Moeda']

#Adicionando a quantidade de moedas na coluna 3
for lin in range(0, tam_moeda):
    pag_inicial.cell(row=lin+tam_acao+3, column=3).value = ls_moeda[lin]['Quantidade de Moeda']

#Adicionando o valor das moedas na coluna 4
for lin in range(0, tam_moeda):
    pag_inicial.cell(row=lin+tam_acao+3, column=4).value = ls_moeda[lin]['Valor da Moeda']

#Adicionando o valor investido no ativo na coluna 5
for lin in range(0, tam_moeda):
    pag_inicial.cell(row=lin+tam_acao+3, column=5).value = ls_moeda[lin]['Valor Investido']

#Adicionando o tipo moeda na coluna 6
for lin in range(0, tam_moeda):
    pag_inicial.cell(row=lin+tam_acao+3, column=6).value = ls_moeda[lin]['Tipo']

#Criando uma lista com o patrimônio total
ls_patrimonio = ["Patrimonio Total", valor_da_carteira]

#Adicionando a lista patrimônio total no final da tabela
for lin in range(0, len(ls_patrimonio)):
    pag_inicial.cell(row=tam_total+4,column=lin+2).value = ls_patrimonio[lin]
    
#Ajustando o formato numérico das células da coluna 4
for lin in range(0, tam_total):
    pag_inicial.cell(row=lin+3,column=4).number_format = "R$ #,##0.00"
    
#Ajustando o formato numérico das células da coluna 5
for lin in range(0, tam_total):
    pag_inicial.cell(row=lin+3,column=5).number_format = "R$ #,##0.00"

#Ajustando o formato numérico da célula valor do patrimônio total
pag_inicial.cell(row=tam_total+4,column=3).number_format = "R$ #,##0.00"

#Ajustando o alinhamento das células
for lin in range(0, tam_total + 4):
    for col in range(0, 6):
        pag_inicial.cell(row=lin+2, column=col+2).alignment = Alignment(horizontal="center")

#Ajustando a dimensão das colunas
for col in pag_inicial.columns:
     max_length = 0
     column = col[0].column_letter 
     for cell in col:
          # Necessário para evitar erro em células vazias
         try:
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length + 2) * 1.05
     pag_inicial.column_dimensions[column].width = adjusted_width

#Adicionando o QRCode na planilha
pag_qrcode = arquivo['Valor da Carteira']
img = Image("QRCode_Secreto.png") 
img.height = 500
img.width = 500
pag_qrcode.add_image(img, "A1") 

#Salvar a planilha
x = randint(0,100)
arquivo.save(f'Investimento{x}.xlsx')