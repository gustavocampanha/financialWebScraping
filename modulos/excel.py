#Importando as bibliotecas necessárias para criarmos a função
import pandas as pd
import matplotlib.pyplot as plt
import qrcode
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Border, Side, Color, Font, PatternFill

#Criando o primeiro gráfico
def graf1(dicionario_ativos):

    #Criando uma lista para adicionarmos os ativos do dicionario
    ls_ativos = list()
    #Criando uma lista para adicionarmos o valor de cada ativo do dicionario
    ls_valor_investido = list()

    #Adicionando os dados nas listas
    for ativos in dicionario_ativos:
        ls_ativos.append(ativos)
        ls_valor_investido.append(dicionario_ativos[ativos]['Valor Investido'])

    #Ajustando a sintaxe
    data = {
        'Ativos': ls_ativos,
        'Valor Investido': ls_valor_investido
    }

    #Utilizando o Pandas para criar o gráfico
    df = pd.DataFrame(data,columns=['Ativos','Valor Investido'])
    df.plot(x ='Ativos', y='Valor Investido', kind = 'bar')
    plt.tight_layout()

    #Salvando o gráfico criado
    plt.savefig('grafico_1.png')
    plt.close()

#Criando o segundo grafico
def graf2(dicionario_ativos):

    valor_total_moeda = 0
    valor_total_acao = 0

    #Fazendo o somatório dos valores investidos em cada tipo de investimento
    for ativo in dicionario_ativos:
        if dicionario_ativos[ativo]['Tipo'] == 'Ação':
            valor_total_acao += dicionario_ativos[ativo]['Valor Investido']
        if dicionario_ativos[ativo]['Tipo'] == 'Moeda':
            valor_total_moeda += dicionario_ativos[ativo]['Valor Investido']

    #Ajustando as casas decimais dos valores obtidos
    valor_total_acao = round(valor_total_acao, 2)
    valor_total_moeda = round(valor_total_moeda, 2)

    data = {
        'Moeda': valor_total_moeda,
        'Ação': valor_total_acao
    }

    plt.pie(data.values(), labels=data.keys(), autopct='%1.1f%%')
    plt.title('Disposição de Ativos')

    #Salvando o gráfico
    plt.savefig('grafico_2.png')
    plt.close()

#Criando o terceiro gráfico
def graf3(dicionario_ativos):

    #Colocando as currencies sem repeteção para sabermos quais cambios temos na cartieira
    currencies = set([dicionario_ativos[ativo]['Currency'] for ativo in dicionario_ativos])
    plot_info = {}

    for currency in currencies:

        filtered_currency = { 
            k: v for k, v in dicionario_ativos.items() 
            if v['Currency'] == currency 
        }
        
        #Ajustando os valores para fazermos o somatório
        total = [ dicionario_ativos[elem]['Valor Investido'] for elem in filtered_currency ]
        total = sum(total)
        plot_info[currency] = total

    #Passando as informações para a criação do gráfico
    plt.pie(plot_info.values(), labels=plot_info.keys(), autopct='%1.1f%%')
    plt.title('Distribuição Cambial')

    #Salvando o terceiro gráfico
    plt.savefig('grafico_3.png')
    plt.close()

#Criando o QRCode
def qr_generator(dicionario_ativos):

    #Descobrindo o patrimonio total da pessoa
    valor_da_carteira = 0
    for dicionario in dicionario_ativos:
        valor_da_carteira += dicionario_ativos[dicionario]['Valor Investido']

    #Criando o QRCode
    qr = qrcode.QRCode(
            version=1,
            box_size=10,
            border=5)

    #Ajustando as casas decimais do número
    valor_da_carteira = round(valor_da_carteira,2)

    #Adicionando o "R$"
    valor_final = (f'R$ {valor_da_carteira}')

    qr.add_data(valor_final)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')

    #Salvando o QRCode
    img.save('QRCode_Secreto.png')

#Parte da tabela do Excel
def excel_tabela(dicionario_ativos):

    #Criar uma planilha chamada book
    arquivo = openpyxl.Workbook()
    ws = arquivo.active
    ws.title = "Tabela"

    #Descobrindo o patrimonio total da pessoa
    valor_da_carteira = 0
    for ativos in dicionario_ativos:
        valor_investido = dicionario_ativos[ativos]['Valor Investido']
        valor_da_carteira += valor_investido
    valor_da_carteira = round(valor_da_carteira, 2)

    #Criando outras páginas
    arquivo.create_sheet('Grafico 1')
    arquivo.create_sheet('Grafico 2')
    arquivo.create_sheet('Grafico 3')
    arquivo.create_sheet('Valor da Carteira')

    #Selecionar uma página
    pag_inicial = arquivo['Tabela']

    #Criando a lista do header
    header = ["Nome", "Quantidade", "Valor Unitário", "Valor Investido", "Tipo"]

    #Adicionando o header na tabela
    for col in range(0, 5):
        pag_inicial.cell(row=2,column=col+2).value = header[col]

    #Descobrir a quantidade de ativos na carteira
    ativos_total = len(dicionario_ativos)

    #Adicionando os nomes na coluna 2
    lin = 0
    for ativo in dicionario_ativos:
        pag_inicial.cell(row=lin+3, column=2).value = dicionario_ativos[ativo]['Nome do Ativo']
        lin += 1

    #Adicionando as quantidades de cada ativo na coluna 3
    lin = 0
    for ativo in dicionario_ativos:
        pag_inicial.cell(row=lin+3, column=3).value = dicionario_ativos[ativo]['Quantidade do Ativo']
        lin += 1

    #Adicionando o valor da ação na coluna 4
    lin = 0
    for ativo in dicionario_ativos:
        pag_inicial.cell(row=lin+3, column=4).value = dicionario_ativos[ativo]['Valor do Ativo']
        lin += 1

    #Adicionando o valor investido no ativo na coluna 5
    lin = 0
    for ativo in dicionario_ativos:
        pag_inicial.cell(row=lin+3, column=5).value = dicionario_ativos[ativo]['Valor Investido']
        lin += 1

    #Adicionando o tipo de ativo ação na coluna 6
    lin = 0
    for ativo in dicionario_ativos:
        pag_inicial.cell(row=lin+3, column=6).value = dicionario_ativos[ativo]['Tipo']
        lin += 1

    #Criando uma lista com o patrimônio total
    ls_patrimonio = ["Patrimonio Total", valor_da_carteira]

    #Adicionando a lista patrimônio total no final da tabela
    for lin in range(0, len(ls_patrimonio)):
        pag_inicial.cell(row=ativos_total+4,column=lin+2).value = ls_patrimonio[lin]
        
    #Ajustando o formato numérico das células da coluna 4
    for lin in range(0, ativos_total):
        pag_inicial.cell(row=lin+3,column=4).number_format = "R$ #,##0.00"
        
    #Ajustando o formato numérico das células da coluna 5
    for lin in range(0, ativos_total):
        pag_inicial.cell(row=lin+3,column=5).number_format = "R$ #,##0.00"

    #Ajustando o formato numérico da célula valor do patrimônio total
    pag_inicial.cell(row=ativos_total+4,column=3).number_format = "R$ #,##0.00"

    #Ajustando o alinhamento das células
    for lin in range(0, ativos_total + 4):
        for col in range(0, 6):
            pag_inicial.cell(row=lin+2, column=col+2).alignment = Alignment(horizontal="center")

    #Ajustando a dimensão das colunas
    for col in pag_inicial.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            # Necessário para evitar erro em células vazias
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.05
        pag_inicial.column_dimensions[column].width = adjusted_width

    #Adicionando borda na tabela
    border_header = Side(border_style='medium', color="000000")
    border_sheet = Side(border_style='thin', color="000000")
    border_header = Border(top=border_header, bottom=border_header, right=border_header, left=border_header)
    border_sheet = Border(top=border_sheet, bottom=border_sheet, right=border_sheet, left=border_sheet)

    #Adicionando borda nas células com os dados do patrimônio
    for lin in range(0, ativos_total):
        for col in range(0, 5):
            pag_inicial.cell(row=lin+3, column=col +2).border = border_sheet
    for col in range(0,5):
        pag_inicial.cell(row=2, column=col +2).border = border_header

    #Adicionando borda nas células do patrimônio
    pag_inicial.cell(row=ativos_total+4,column=2).border = border_sheet
    pag_inicial.cell(row=ativos_total+4,column=3).border = border_sheet

    #Criando a cor do header
    headerFill = PatternFill(start_color='4EEE94',
                   end_color='4EEE94',
                   fill_type='solid')

    #Adicionando preenchimento no header
    for a in range(0, 5):
        pag_inicial.cell(row=2, column=a+2).fill = headerFill

    #Adicionando o gráfico 1
    graf1 = arquivo['Grafico 1']
    imgg1 = Image("grafico_1.png") 
    graf1.add_image(imgg1, "A1")

    #Adicionando o gráfico 2
    graf2 = arquivo['Grafico 2']
    imgg2 = Image("grafico_2.png") 
    graf2.add_image(imgg2, "A1")

    #Adicionando o gráfico 3
    graf3 = arquivo['Grafico 3']
    imgg3 = Image("grafico_3.png") 
    graf3.add_image(imgg3, "A1")

    #Adicionando o QRCode na planilha
    pag_qrcode = arquivo['Valor da Carteira']
    imgqr = Image("QRCode_Secreto.png") 
    imgqr.height = 500
    imgqr.width = 500
    pag_qrcode.add_image(imgqr, "A1")

    #Retirando as linhas de grade
    pag_inicial.sheet_view.showGridLines = False
    graf1.sheet_view.showGridLines = False
    graf2.sheet_view.showGridLines = False
    graf3.sheet_view.showGridLines = False
    pag_qrcode.sheet_view.showGridLines = False

    #Salvar a planilha
    arquivo.save(f'planilhaInvestimento.xlsx')